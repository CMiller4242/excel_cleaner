#!/usr/bin/env python3
"""
Test with raw Excel data using openpyxl to check for cell-level issues
"""

import pandas as pd
import openpyxl
import sys
from pathlib import Path

# Add parent to path
sys.path.insert(0, str(Path(__file__).parent))

from engine.executor import OperationExecutor

def test_with_openpyxl():
    """Use openpyxl to read raw cell data"""

    filename = "Volunteer Directors 11.20.25(2).xlsx"

    print("=" * 100)
    print(f"RAW EXCEL ANALYSIS: {filename}")
    print("=" * 100)

    # First, load with pandas (normal way)
    print(f"\n[1] Loading with pandas...")
    df = pd.read_excel(filename)
    person_street_pd = df['Person Street']
    print(f"    Pandas sees:")
    print(f"      Total: {len(person_street_pd)}")
    print(f"      Non-blank: {person_street_pd.notna().sum()}")
    print(f"      Blank: {person_street_pd.isna().sum()}")

    # Now load with openpyxl to see raw cell data
    print(f"\n[2] Loading with openpyxl (raw cell data)...")
    wb = openpyxl.load_workbook(filename, data_only=True)
    ws = wb.active

    # Find Person Street column
    header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    person_street_col_idx = None
    for idx, col_name in enumerate(header_row, start=1):
        if col_name == 'Person Street':
            person_street_col_idx = idx
            break

    if person_street_col_idx is None:
        print("    ERROR: Person Street column not found!")
        return

    print(f"    Person Street is column {person_street_col_idx}")

    # Read all Person Street values from Excel
    person_street_cells = []
    for row_idx in range(2, ws.max_row + 1):  # Skip header
        cell = ws.cell(row=row_idx, column=person_street_col_idx)
        person_street_cells.append({
            'row': row_idx,
            'value': cell.value,
            'type': type(cell.value).__name__,
            'is_none': cell.value is None
        })

    # Count blanks in Excel
    excel_blanks = sum(1 for c in person_street_cells if c['value'] is None)
    excel_non_blanks = len(person_street_cells) - excel_blanks

    print(f"    Openpyxl sees:")
    print(f"      Total: {len(person_street_cells)}")
    print(f"      Non-blank: {excel_non_blanks}")
    print(f"      Blank (None): {excel_blanks}")

    # Compare pandas vs openpyxl
    print(f"\n[3] Comparison:")
    if person_street_pd.notna().sum() == excel_non_blanks:
        print(f"    ✓ Pandas and openpyxl agree on blank count")
    else:
        print(f"    ⚠️  MISMATCH!")
        print(f"       Pandas non-blank: {person_street_pd.notna().sum()}")
        print(f"       Openpyxl non-blank: {excel_non_blanks}")
        print(f"       Difference: {abs(person_street_pd.notna().sum() - excel_non_blanks)}")

    # Show some non-blank values from Excel
    print(f"\n[4] First 10 non-blank values from Excel (openpyxl):")
    non_blanks = [c for c in person_street_cells if c['value'] is not None][:10]
    for cell in non_blanks:
        print(f"      Row {cell['row']}: '{cell['value']}' (type: {cell['type']})")

    # Now run the operation and see what gets removed
    print(f"\n[5] Running Remove Rows If operation...")
    operations = [
        {
            'operation_id': 'data_remove_rows_if',
            'parameters': {
                'column': 'Person Street',
                'condition': 'is_blank',
                'enhanced_blank_detection': False
            },
            'enabled': True
        }
    ]

    executor = OperationExecutor()
    result_df, removed_df = executor.execute_queue_with_tracking(df.copy(), operations)

    print(f"    Results: {len(result_df)} rows")
    print(f"    Removed: {len(removed_df)} rows")

    # Check what's in removed
    if 'Person Street' in removed_df.columns:
        removed_ps = removed_df['Person Street']
        non_blank_removed = removed_ps.notna().sum()

        print(f"\n[6] Analysis of REMOVED sheet:")
        print(f"    Total removed: {len(removed_df)}")
        print(f"    Non-blank Person Street in removed: {non_blank_removed}")

        if non_blank_removed > 0:
            print(f"\n    ⚠️  FALSE POSITIVES: {non_blank_removed} non-blank values removed!")
            print(f"\n    First 20 false positives:")
            for idx, val in removed_df[removed_ps.notna()].head(20).iterrows():
                addr = val['Person Street']
                print(f"      [{idx}] '{addr}'")
        else:
            print(f"    ✓ All removed values are blank (correct)")

    wb.close()

if __name__ == '__main__':
    test_with_openpyxl()
