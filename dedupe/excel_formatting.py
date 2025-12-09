"""
Professional Excel Formatting Module
Provides styling functions for comparison mode output sheets

IMPORTANT: This module ONLY applies styles to existing data.
DO NOT write any Excel formulas (e.g., ws["B5"] = "=SUM(B6:B10)").
All values must be pre-computed in Python and written as plain values.
This prevents Excel XML corruption issues.
"""

from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, NamedStyle
from openpyxl.utils import get_column_letter


# Define color constants
COLOR_BLUE_HEADER = "1F4E79"
COLOR_GRAY_HEADER = "404040"
COLOR_RED_HEADER = "C00000"
COLOR_ORANGE_HEADER = "F4B084"
COLOR_LIGHT_GRAY = "F2F2F2"
COLOR_LIGHT_YELLOW = "FFF2CC"
COLOR_WHITE = "FFFFFF"


def create_named_styles(workbook):
    """Create reusable named styles for the workbook"""

    # Header style (dark blue with white text)
    if 'header_style' not in workbook.named_styles:
        header_style = NamedStyle(name='header_style')
        header_style.font = Font(name='Segoe UI', size=11, bold=True, color=COLOR_WHITE)
        header_style.fill = PatternFill(start_color=COLOR_BLUE_HEADER, end_color=COLOR_BLUE_HEADER, fill_type='solid')
        header_style.alignment = Alignment(horizontal='center', vertical='center')
        header_style.border = Border(
            left=Side(style='thin', color=COLOR_WHITE),
            right=Side(style='thin', color=COLOR_WHITE),
            top=Side(style='thin', color=COLOR_WHITE),
            bottom=Side(style='thin', color=COLOR_WHITE)
        )
        workbook.add_named_style(header_style)

    # Gray header style (for raw sheets)
    if 'gray_header_style' not in workbook.named_styles:
        gray_header_style = NamedStyle(name='gray_header_style')
        gray_header_style.font = Font(name='Segoe UI', size=11, bold=True, color=COLOR_WHITE)
        gray_header_style.fill = PatternFill(start_color=COLOR_GRAY_HEADER, end_color=COLOR_GRAY_HEADER, fill_type='solid')
        gray_header_style.alignment = Alignment(horizontal='center', vertical='center')
        gray_header_style.border = Border(
            left=Side(style='thin', color=COLOR_WHITE),
            right=Side(style='thin', color=COLOR_WHITE),
            top=Side(style='thin', color=COLOR_WHITE),
            bottom=Side(style='thin', color=COLOR_WHITE)
        )
        workbook.add_named_style(gray_header_style)

    # Red header style (for overlapping records)
    if 'red_header_style' not in workbook.named_styles:
        red_header_style = NamedStyle(name='red_header_style')
        red_header_style.font = Font(name='Segoe UI', size=11, bold=True, color=COLOR_WHITE)
        red_header_style.fill = PatternFill(start_color=COLOR_RED_HEADER, end_color=COLOR_RED_HEADER, fill_type='solid')
        red_header_style.alignment = Alignment(horizontal='center', vertical='center')
        red_header_style.border = Border(
            left=Side(style='thin', color=COLOR_WHITE),
            right=Side(style='thin', color=COLOR_WHITE),
            top=Side(style='thin', color=COLOR_WHITE),
            bottom=Side(style='thin', color=COLOR_WHITE)
        )
        workbook.add_named_style(red_header_style)

    # Data even row style (white background)
    if 'data_even_style' not in workbook.named_styles:
        data_even_style = NamedStyle(name='data_even_style')
        data_even_style.font = Font(name='Segoe UI', size=10)
        data_even_style.fill = PatternFill(start_color=COLOR_WHITE, end_color=COLOR_WHITE, fill_type='solid')
        data_even_style.alignment = Alignment(horizontal='left', vertical='center')
        workbook.add_named_style(data_even_style)

    # Data odd row style (light gray background)
    if 'data_odd_style' not in workbook.named_styles:
        data_odd_style = NamedStyle(name='data_odd_style')
        data_odd_style.font = Font(name='Segoe UI', size=10)
        data_odd_style.fill = PatternFill(start_color=COLOR_LIGHT_GRAY, end_color=COLOR_LIGHT_GRAY, fill_type='solid')
        data_odd_style.alignment = Alignment(horizontal='left', vertical='center')
        workbook.add_named_style(data_odd_style)

    # Metric label style (for summary report)
    if 'metric_label_style' not in workbook.named_styles:
        metric_label_style = NamedStyle(name='metric_label_style')
        metric_label_style.font = Font(name='Segoe UI', size=10, bold=True)
        metric_label_style.alignment = Alignment(horizontal='left', vertical='center')
        workbook.add_named_style(metric_label_style)

    # Metric value style (for summary report)
    if 'metric_value_style' not in workbook.named_styles:
        metric_value_style = NamedStyle(name='metric_value_style')
        metric_value_style.font = Font(name='Segoe UI', size=10)
        metric_value_style.alignment = Alignment(horizontal='right', vertical='center')
        workbook.add_named_style(metric_value_style)

    # Highlight total style (yellow background, bold)
    if 'highlight_total_style' not in workbook.named_styles:
        highlight_total_style = NamedStyle(name='highlight_total_style')
        highlight_total_style.font = Font(name='Segoe UI', size=10, bold=True)
        highlight_total_style.fill = PatternFill(start_color=COLOR_LIGHT_YELLOW, end_color=COLOR_LIGHT_YELLOW, fill_type='solid')
        highlight_total_style.alignment = Alignment(horizontal='left', vertical='center')
        highlight_total_style.border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        workbook.add_named_style(highlight_total_style)


def apply_header_formatting(worksheet, header_style='header_style'):
    """Apply formatting to header row"""
    for cell in worksheet[1]:
        cell.style = header_style
    worksheet.row_dimensions[1].height = 22


def apply_alternating_rows(worksheet, start_row=2):
    """Apply alternating row colors to data rows"""
    max_row = worksheet.max_row
    max_col = worksheet.max_column

    for row_idx in range(start_row, max_row + 1):
        style = 'data_even_style' if row_idx % 2 == 0 else 'data_odd_style'
        for col_idx in range(1, max_col + 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            cell.style = style
            worksheet.row_dimensions[row_idx].height = 16


def autofit_columns(worksheet):
    """Auto-fit column widths based on content"""
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)

        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass

        # Set width with some padding (max 50 to prevent excessive widths)
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width


def freeze_header_row(worksheet):
    """Freeze the header row"""
    worksheet.freeze_panes = "A2"


def hide_gridlines(worksheet):
    """Hide gridlines in worksheet view"""
    worksheet.sheet_view.showGridLines = False


def format_summary_report(worksheet):
    """Format Summary_Report sheet with special styling"""
    # Apply blue header
    apply_header_formatting(worksheet, 'header_style')

    # Apply metric label style to column A
    for row_idx in range(2, worksheet.max_row + 1):
        cell = worksheet.cell(row=row_idx, column=1)

        # Safety check: ensure cell value doesn't start with '=' (would be interpreted as formula)
        # This should already be handled by clean_excel_text(), but double-check here
        if cell.value and isinstance(cell.value, str) and cell.value.startswith('=') and not cell.value.startswith("'="):
            # Prefix with apostrophe to force text interpretation
            cell.value = "'" + cell.value

        cell.style = 'metric_label_style'

        # Apply metric value style to column B
        value_cell = worksheet.cell(row=row_idx, column=2)
        value_cell.style = 'metric_value_style'

        # Highlight key totals (rows containing "Overlapping", "Combined Unique", "Overlap Rate")
        if cell.value and any(keyword in str(cell.value) for keyword in ['Overlapping', 'Combined Unique', 'Overlap Rate']):
            cell.style = 'highlight_total_style'
            value_cell.style = 'highlight_total_style'

    # Auto-fit columns
    autofit_columns(worksheet)

    # Freeze header
    freeze_header_row(worksheet)

    # Hide gridlines
    hide_gridlines(worksheet)


def format_data_table(worksheet, header_style='header_style'):
    """Format a standard data table (Combined_Cleaned, Raw sheets)"""
    # Apply header formatting
    apply_header_formatting(worksheet, header_style)

    # Apply alternating rows
    apply_alternating_rows(worksheet, start_row=2)

    # Auto-fit columns
    autofit_columns(worksheet)

    # Freeze header
    freeze_header_row(worksheet)

    # Hide gridlines
    hide_gridlines(worksheet)


def format_overlapping_records(worksheet):
    """Format Overlapping_Records sheet with red/orange theme"""
    # Apply red header
    apply_header_formatting(worksheet, 'red_header_style')

    # Apply alternating rows
    apply_alternating_rows(worksheet, start_row=2)

    # Bold Match_Reason column if it exists
    header_row = [cell.value for cell in worksheet[1]]
    if 'Match_Reason' in header_row:
        match_reason_col = header_row.index('Match_Reason') + 1
        for row_idx in range(2, worksheet.max_row + 1):
            cell = worksheet.cell(row=row_idx, column=match_reason_col)
            cell.font = Font(name='Segoe UI', size=10, bold=True)

    # Auto-fit columns
    autofit_columns(worksheet)

    # Freeze header
    freeze_header_row(worksheet)

    # Hide gridlines
    hide_gridlines(worksheet)


def apply_professional_formatting(workbook):
    """
    Apply professional formatting to all sheets in the workbook

    Args:
        workbook: openpyxl Workbook object
    """
    # Create named styles first
    create_named_styles(workbook)

    # Format each sheet based on its name
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]

        if sheet_name == 'Summary_Report':
            format_summary_report(worksheet)
        elif sheet_name == 'Combined_Cleaned':
            format_data_table(worksheet, 'header_style')
        elif sheet_name in ['File1_Raw', 'File2_Raw']:
            format_data_table(worksheet, 'gray_header_style')
        elif sheet_name == 'Overlapping_Records':
            format_overlapping_records(worksheet)
        else:
            # Default formatting for any other sheets
            format_data_table(worksheet, 'header_style')
